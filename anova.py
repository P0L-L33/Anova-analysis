from openpyxl import Workbook
from openpyxl import load_workbook
import statistics as stats

#convention: for saving the averages of each sample group, we assume that we don't have more than 7 sample groups, and place them in column G

def calculate_gr_average():
    global column, values, hoho, SSW, SSB, averages, n

    total = 0
    values = []
    colC = ws[column]
    for loners in colC:
        print(loners.value)
        values.append(loners.value)
    hoho = stats.mean(values)
    n = n + len(values)
    lengths.append(len(values))
    averages.append(hoho)


    #calculate_SSW()
    for j in values:
        SSW = SSW + (j - hoho) ** 2

    #calculade_SSB()


    print(hoho)
    ws.cell(row=nr, column=7).value = hoho

    print("Input a column name please:")
    column = input()

#def calculate_SSW():
    #for j in values:
        #ws["I2"] = ws["I2"] + (j - hoho)^2




wb = load_workbook('anova.xlsx')
ws = wb.active
averages = []
lengths = []
SSW = 0
SSB = 0
print("Input a column name please:")
column = input()
nr = 0
n = 0
while column != '0': #convention to end the loop
    nr = nr + 1
    calculate_gr_average()
ws.cell(row=nr+1, column=7).value = stats.mean(averages)
K = n - len(averages)

colG = ws['G']
for b in range(3):
    SSB = SSB + lengths[b]*(colG[b+1].value - stats.mean(averages))**2
print(SSB)

ws['H1'] = "Source of Variation"
ws['H2'] = "Between Groups"
ws['H3'] = "Within Groups"
ws['H4'] = "Total"
ws['I1'] = "SS"
ws['J1'] = "df"
ws['K1'] = "MS variance"
ws['L1'] = "F ratio"
ws['J2'] = K - 1
ws['J3'] = n - K
ws['J4'] = n - 1

ws['I3'] = SSW


wb.save('anova.xlsx')